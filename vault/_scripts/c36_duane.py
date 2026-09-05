#!/usr/bin/env python3
"""c36_duane.py — a Crow-AMSAA reliability-growth exponent for fishery-management programmes.

Computes the Duane/Crow growth exponent beta for regional fishery-management
programmes, treating each assessed stock-year as one unit of cumulative programme
operating time and each stock-year that breaches a reference point as one failure
of a repairable system.

    NHPP power-law intensity   lambda(t) = lambda * beta * t^(beta-1)
    time-truncated MLE         beta_hat = n / sum_i ln(T / t_i)
                               lambda_hat = n / T^beta_hat
    exact CI (Crow 1974/1982)  beta_hat * chi2_{a/2,2n}/(2n) , beta_hat * chi2_{1-a/2,2n}/(2n)

beta < 1  failure intensity falling  = the programme is learning
beta = 1  homogeneous Poisson        = no learning
beta > 1  failure intensity rising   = the programme is getting worse

Data: RAM Legacy Stock Assessment Database v4.65, Ricard et al., Zenodo record
11995054, DOI 10.5281/zenodo.11995054 (deposited 2024-06-17). The script reads
the distributed zip directly; pass --zip PATH, or let it download to --cache.

    python c36_duane.py --zip "RAMLDB v4.65.zip"

Dependencies: stdlib + openpyxl + scipy (chi2 quantiles only) + no pandas.
Run 2026-09-05 for vault/computed/C36-conservation-duane.md.
"""
import argparse, collections, csv, io, math, os, random, sys, urllib.request, zipfile

ZENODO = ("https://zenodo.org/api/records/11995054/files/"
          "RAMLDB%20v4.65.zip/content")
XLSX = "Excel/RAMLDB v4.65 (assessment data only).xlsx"

# --------------------------------------------------------------------------
# Crow-AMSAA


def crow(times, T):
    """Time-truncated Crow-AMSAA MLE. `times` are cumulative-exposure failure
    epochs in (0, T]. Returns (beta, lo95, hi95, lambda, n) or None."""
    from scipy.stats import chi2
    n = len(times)
    if n < 5 or T <= 0:
        return None
    s = sum(math.log(T / t) for t in times)
    if s <= 0:
        return None
    beta = n / s
    lo = beta * chi2.ppf(0.025, 2 * n) / (2 * n)
    hi = beta * chi2.ppf(0.975, 2 * n) / (2 * n)
    return beta, lo, hi, n / T ** beta, n


# --------------------------------------------------------------------------
# Data


def load(zpath):
    """-> (rows, region) where rows = [(stockid, year, UdivUmsypref,
    BdivBmsypref)] and region maps stockid -> RAM region string."""
    import openpyxl
    with zipfile.ZipFile(zpath) as z:
        buf = io.BytesIO(z.read(XLSX))
    wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)

    ws = wb["stock"]
    it = ws.iter_rows(values_only=True)
    h = list(next(it))
    i_id, i_rg = h.index("stockid"), h.index("region")
    region = {r[i_id]: r[i_rg] for r in it if r[i_id]}

    ws = wb["timeseries_values_views"]
    it = ws.iter_rows(values_only=True)
    h = list(next(it))
    ix = {k: h.index(k) for k in
          ("stockid", "year", "UdivUmsypref", "BdivBmsypref")}
    rows = []
    for r in it:
        sid = r[ix["stockid"]]
        if not sid:
            continue
        try:
            year = int(float(r[ix["year"]]))
        except (TypeError, ValueError):
            continue

        def num(k):
            try:
                return float(r[ix[k]])
            except (TypeError, ValueError):
                return None
        u, b = num("UdivUmsypref"), num("BdivBmsypref")
        if u is None and b is None:
            continue
        rows.append((sid, year, u, b))
    return rows, region


def failed(row, mode):
    """mode 'U': overfishing, U/Umsy > 1. mode 'B': overfished, B/Bmsy < 0.5."""
    return row[2] > 1.0 if mode == "U" else row[3] < 0.5


def sequence(rows, region, rg, mode, y0, y1, balanced):
    """Order the region's stock-years by (year, stockid) and return
    (failure epochs, T, n_stocks). `balanced` keeps only stocks present in
    EVERY year of [y0, y1], so exposure per calendar year is constant."""
    k = 2 if mode == "U" else 3
    d = [r for r in rows if region.get(r[0]) == rg
         and y0 <= r[1] <= y1 and r[k] is not None]
    if balanced:
        need = set(range(y0, y1 + 1))
        seen = collections.defaultdict(set)
        for r in d:
            seen[r[0]].add(r[1])
        keep = {s for s, ys in seen.items() if need <= ys}
        d = [r for r in d if r[0] in keep]
    d.sort(key=lambda r: (r[1], r[0]))
    epochs = [i for i, r in enumerate(d, 1) if failed(r, mode)]
    return epochs, len(d), len({r[0] for r in d})


# --------------------------------------------------------------------------
# Controls


def null_control(T, n, reps=2000, seed=7):
    """Homogeneous Poisson null: n failures placed uniformly over T slots.
    An unbiased estimator must return beta ~ 1 here."""
    random.seed(seed)
    bs = []
    for _ in range(reps):
        e = sorted(random.sample(range(1, T + 1), n))
        r = crow(e, T)
        if r:
            bs.append(r[0])
    bs.sort()
    q = lambda p: bs[int(p * (len(bs) - 1))]
    return q(0.5), q(0.025), q(0.975)


def recovery(beta_true, T, n_target, seed=7):
    """Simulate a power-law NHPP thinned onto T integer slots and re-fit.
    Quantifies the discretisation bias at small beta."""
    random.seed(seed)
    lam = n_target / T ** beta_true
    e = [i for i in range(1, T + 1)
         if random.random() < min(lam * (i ** beta_true - (i - 1) ** beta_true), 1)]
    return crow(e, T)


# --------------------------------------------------------------------------


def main(argv=None):
    ap = argparse.ArgumentParser()
    ap.add_argument("--zip", default="RAMLDB v4.65.zip")
    ap.add_argument("--cache", default=".")
    ap.add_argument("--window", default="1990,2015")
    args = ap.parse_args(argv)

    zpath = args.zip
    if not os.path.exists(zpath):
        zpath = os.path.join(args.cache, "RAMLDB v4.65.zip")
        if not os.path.exists(zpath):
            print("downloading RAM Legacy v4.65 (117 MB) ...", file=sys.stderr)
            req = urllib.request.Request(
                ZENODO, headers={"User-Agent":
                                 "biomimicry/1.0 (mailto:deciduusleaf@gmail.com)"})
            with urllib.request.urlopen(req, timeout=900) as r, open(zpath, "wb") as f:
                f.write(r.read())

    y0, y1 = (int(x) for x in args.window.split(","))
    rows, region = load(zpath)
    regions = sorted({v for v in region.values() if v})

    for balanced in (False, True):
        for mode in ("U", "B"):
            tag = "BALANCED" if balanced else "ALL STOCK-YEARS"
            crit = "U/Umsy > 1" if mode == "U" else "B/Bmsy < 0.5"
            print(f"\n=== {tag}  {y0}-{y1}  failure = {crit}")
            print(f"{'region':30s} {'stocks':>6} {'T':>6} {'n':>6} "
                  f"{'beta':>7} {'lo95':>7} {'hi95':>7}")
            for rg in regions:
                e, T, k = sequence(rows, region, rg, mode,
                                   y0 if balanced else 1950, y1, balanced)
                r = crow(e, T)
                if not r:
                    continue
                print(f"{rg:30s} {k:6d} {T:6d} {r[4]:6d} "
                      f"{r[0]:7.3f} {r[1]:7.3f} {r[2]:7.3f}")

    print("\n=== CONTROL 1: homogeneous-Poisson null (estimator must give beta ~ 1)")
    for T, n in ((806, 233), (1404, 1010), (208, 166), (702, 335)):
        m, lo, hi = null_control(T, n)
        print(f"  T={T:5d} n={n:5d}  beta median {m:.3f}  [{lo:.3f}, {hi:.3f}]")

    print("\n=== CONTROL 2: recovery of a known beta on T=806 integer slots")
    for b in (0.4, 0.5, 0.7, 1.0):
        r = recovery(b, 806, 233)
        print(f"  true {b:.1f} -> {r[0]:.3f} [{r[1]:.3f}, {r[2]:.3f}]  n={r[4]}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
