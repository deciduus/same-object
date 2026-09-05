#!/usr/bin/env python3
"""
c32_replication.py - replicate [[C29-recovery-beta]]'s Weibull recovery beta on an
INDEPENDENT recovery dataset: Moreno-Mateos et al. 2017's recovery-debt database.

WHAT THIS TESTS
---------------
C29 fitted Jones & Schmitz 2009 Table S1 as a right-censored survival problem and got
pooled beta = 0.587 [0.510, 0.668], with a per-habitat ordering.  C29 sec.5 states two
falsifiable predictions:

  P1  Re-fit on Moreno-Mateos 2017 or Crouzeilles 2016.  The SIGN of the Spearman rank
      correlation of per-habitat beta between the two datasets should be POSITIVE, with
      marine/brackish lowest and freshwater/agricultural presses highest.
  P2  Split by RESPONSE-VARIABLE CLASS.  If beta < 1 is real deceleration it survives the
      split within a class; if it is frailty (a mixture of fast and slow recoverers) the
      within-class beta rises toward 1 while the pooled beta stays low.

DATA
----
Moreno-Mateos, D. et al. (2017) "Anthropogenic ecosystem disturbance and the recovery
debt", Nature Communications 8:14163, doi 10.1038/ncomms14163.  The paper's own Data
Availability statement points at Dryad doi:10.5061/dryad.t5c97, file
"Moreno, Jones database.xlsx" (400,066 bytes, 3,816 outcome-measure rows, 356 primary
studies, 26 columns).  The Nature Communications supplement itself is figures + summary
tables + a reference list only -- there is NO machine-readable per-study table there.

NOTE ON FETCHING.  datadryad.org sits behind a proof-of-work bot check (Anubis) which
this script does NOT attempt to defeat.  Download the file by hand in a browser from
https://datadryad.org/dataset/doi:10.5061/dryad.t5c97 and pass --xlsx.  Fetched
2026-09-05 for the numbers in [[C32-recovery-beta-replication]].

Crouzeilles et al. 2016 (doi 10.1038/ncomms11666, Dryad doi:10.5061/dryad.k3479) was
checked as the alternative target and REJECTED: its `Meta_analysis.txt` carries
site / disturbance / taxon-group dummies and a log response ratio `RR`, but NO time
column and no recovery yes/no, so no survival time can be constructed from it at all.

SURVIVAL CODING (the whole methodological content of this script)
-----------------------------------------------------------------
Moreno-Mateos measures a recovery DEBT, not a recovery TIME.  Each row is one outcome
measure with:

    Start (Xs)  value when recovery started
    End   (Xe)  value at the end of the observation window
    Goal  (Xr)  value in the reference / undisturbed system
    Time since restoration started (T), in years

The survival variable is constructed as:

    recovered  iff  Xe has reached or crossed Xr in the direction of travel from Xs
                    (Xs < Xr  ->  Xe >= Xr ;   Xs > Xr  ->  Xe <= Xr)
    time       =    T

    "recovered"     -> EVENT     at T   (the C29 coding)
    "not recovered" -> RIGHT-CENSORED at T

This is the same yes/no-plus-a-duration structure Jones & Schmitz tabulate directly, and
it is what makes the two datasets comparable at all.  It costs one assumption that C29
did not have to make, and the note says so: **a row coded as an event at T reached its
goal at some unknown time <= T**, because Moreno-Mateos observed status at T rather than
recording a return time.  Coding the event AT T therefore OVERSTATES recovery times for
the recovered rows.  Run `--likelihood current-status` for the estimator that handles
this correctly (each row contributes log F(T) if recovered, log S(T) if not); the
headline uses `--likelihood c29` because P1 is a comparison of estimators, not of data.

Rows with Xs == Xr (already at the reference at t=0), T <= 0, or a missing field are
dropped and counted.

HABITAT MAPPING (declared, not fitted)
--------------------------------------
C29 habitat            Moreno-Mateos "Habitat category"
  Forest          <-   Forest
  Marine          <-   Marine (benthic) + Marine (pelagic)
  Freshwater      <-   Lake + River + Freshwater wetland
  Brackish        <-   Tidal wetland + Mangrove
  Terrestrial     <-   Grassland

DEPENDENCIES:  numpy, scipy, openpyxl, xlrd (xlrd only for --overlap)

USAGE
    python c32_replication.py --xlsx "Moreno, Jones database.xlsx"
    python c32_replication.py --xlsx FILE --likelihood current-status
    python c32_replication.py --xlsx FILE --study-level
    python c32_replication.py --xlsx FILE --overlap --jones-xls jones_s1.xls
"""

import argparse
import collections
import math
import re
import sys

import numpy as np
from scipy.optimize import brentq, minimize_scalar

# ---------------------------------------------------------------- C29's estimator
# Copied verbatim in substance from vault/_scripts/c29_recovery.py so the two notes
# share one estimator.  eta is profiled out analytically; the CI is profile-likelihood
# at a chi2_1 drop of 1.921, not a normal approximation.


def _neg_logL_c29(beta, t, e):
    d = float(e.sum())
    with np.errstate(over="ignore"):
        eta = ((t ** beta).sum() / d) ** (1.0 / beta)
        if not np.isfinite(eta) or eta <= 0:
            return 1e12
        return -(d * np.log(beta) - d * beta * np.log(eta)
                 + (beta - 1.0) * np.log(t[e == 1]).sum()
                 - ((t / eta) ** beta).sum())


def _neg_logL_cs(params, t, e):
    """Current-status (case-1 interval censored) log-likelihood; eta NOT profilable."""
    beta, logeta = params
    if beta <= 0:
        return 1e12
    eta = math.exp(logeta)
    with np.errstate(over="ignore", under="ignore"):
        S = np.exp(-((t / eta) ** beta))
        S = np.clip(S, 1e-12, 1 - 1e-12)
        return -(np.log(1.0 - S[e == 1]).sum() + np.log(S[e == 0]).sum())


def _fit_cs(t, e, level=1.920729):
    from scipy.optimize import minimize

    def prof(beta):
        r = minimize(lambda le: _neg_logL_cs((beta, le[0]), t, e),
                     x0=[math.log(max(np.median(t), 1e-3))], method="Nelder-Mead",
                     options={"xatol": 1e-6, "fatol": 1e-9})
        return r.fun, math.exp(r.x[0])

    r = minimize_scalar(lambda b: prof(b)[0], bounds=(0.02, 25.0), method="bounded")
    beta, l0 = r.x, r.fun
    eta = prof(beta)[1]

    def gap(b):
        return prof(b)[0] - l0 - level

    lo = brentq(gap, 1e-2, beta) if gap(1e-2) > 0 else float("nan")
    hi = brentq(gap, beta, 60.0) if gap(60.0) > 0 else float("nan")
    return beta, lo, hi, eta


def fit_weibull(pairs, likelihood="c29", level=1.920729):
    """(n, n_events, n_censored, beta, lo, hi, eta) for [(time, event), ...]."""
    t = np.asarray([p[0] for p in pairs], float)
    e = np.asarray([p[1] for p in pairs], int)
    if e.sum() < 3 or (len(e) - e.sum()) < 1:
        return None
    if likelihood == "current-status":
        beta, lo, hi, eta = _fit_cs(t, e, level)
        return (len(t), int(e.sum()), len(t) - int(e.sum()), beta, lo, hi, eta)
    r = minimize_scalar(_neg_logL_c29, bounds=(0.02, 25.0), args=(t, e), method="bounded")
    beta, l0 = r.x, r.fun

    def gap(b):
        return _neg_logL_c29(b, t, e) - l0 - level

    lo = brentq(gap, 1e-3, beta) if gap(1e-3) > 0 else float("nan")
    hi = brentq(gap, beta, 60.0) if gap(60.0) > 0 else float("nan")
    eta = ((t ** beta).sum() / e.sum()) ** (1.0 / beta)
    return (len(t), int(e.sum()), len(t) - int(e.sum()), beta, lo, hi, eta)


# ---------------------------------------------------------------- data

HABITAT_MAP = {
    "Forest": "Forest",
    "Marine (benthic)": "Marine",
    "Marine (pelagic)": "Marine",
    "Lake": "Freshwater",
    "River": "Freshwater",
    "Freshwater wetland": "Freshwater",
    "Tidal wetland": "Brackish",
    "Mangrove": "Brackish",
    "Grassland": "Terrestrial",
}

# C29 sec.3, table "beta by habitat" - the numbers P1 is correlated against.
C29_BETA = {"Forest": 0.769, "Marine": 0.644, "Freshwater": 0.893,
            "Brackish": 0.501, "Terrestrial": 0.570}

# Response-variable classes for P2.  Moreno-Mateos "Metric type" values.
METRIC_CLASS = {"Abundance": "structural/biomass", "C": "structural/biomass",
                "N": "structural/biomass", "Organic matter": "structural/biomass",
                "Diversity": "compositional/species"}


def load(xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sh = wb[wb.sheetnames[0]]
    rows = list(sh.iter_rows(values_only=True))
    hdr = {h: i for i, h in enumerate(rows[0])}
    out = []
    for r in rows[1:]:
        if r[hdr["Citation"]] in (None, ""):
            continue
        out.append({
            "citation": str(r[hdr["Citation"]]).strip(),
            "habitat_raw": str(r[hdr["Habitat category"]]).strip(),
            "disturbance": re.sub(r"\s+", " ", str(r[hdr["Disturbance category"]])).strip(),
            "metric": str(r[hdr["Metric type"]]).strip(),
            "T": r[hdr["Time since restoration started"]],
            "goal": r[hdr["Goal"]], "start": r[hdr["Start"]], "end": r[hdr["End"]],
        })
    return out


def survival(rows):
    """-> (records, dropped Counter).  record = dict with time, event, habitat, ..."""
    recs, dropped = [], collections.Counter()
    for r in rows:
        h = HABITAT_MAP.get(r["habitat_raw"])
        if h is None:
            dropped["habitat not mappable to a C29 class"] += 1
            continue
        try:
            t = float(r["T"]); xs = float(r["start"])
            xe = float(r["end"]); xr = float(r["goal"])
        except (TypeError, ValueError):
            dropped["missing/non-numeric T, Start, End or Goal"] += 1
            continue
        if not (t > 0):
            dropped["time since restoration <= 0 or absent"] += 1
            continue
        if xs == xr:
            dropped["Start equals Goal (no recovery to observe)"] += 1
            continue
        ev = 1 if ((xs < xr and xe >= xr) or (xs > xr and xe <= xr)) else 0
        recs.append({"time": t, "event": ev, "habitat": h,
                     "disturbance": r["disturbance"], "metric": r["metric"],
                     "citation": r["citation"]})
    return recs, dropped


def study_level(recs):
    """One record per (citation, habitat): recovered if ANY outcome measure recovered,
    at that study's median observation time.  Kills the pseudo-replication."""
    g = collections.defaultdict(list)
    for r in recs:
        g[(r["citation"], r["habitat"])].append(r)
    out = []
    for (cit, hab), v in g.items():
        out.append({"time": float(np.median([x["time"] for x in v])),
                    "event": 1 if any(x["event"] for x in v) else 0,
                    "habitat": hab, "disturbance": v[0]["disturbance"],
                    "metric": v[0]["metric"], "citation": cit})
    return out


# ---------------------------------------------------------------- Spearman

def spearman(x, y):
    def rank(a):
        order = sorted(range(len(a)), key=lambda i: a[i])
        rk = [0.0] * len(a)
        i = 0
        while i < len(a):
            j = i
            while j + 1 < len(a) and a[order[j + 1]] == a[order[i]]:
                j += 1
            avg = (i + j) / 2.0 + 1
            for k in range(i, j + 1):
                rk[order[k]] = avg
            i = j + 1
        return rk
    rx, ry = rank(x), rank(y)
    n = len(x)
    mx, my = sum(rx) / n, sum(ry) / n
    num = sum((a - mx) * (b - my) for a, b in zip(rx, ry))
    den = math.sqrt(sum((a - mx) ** 2 for a in rx) * sum((b - my) ** 2 for b in ry))
    rho = num / den if den else float("nan")
    # exact two-sided permutation p-value (n is 5 or 6 here; 720 permutations at most)
    import itertools
    cnt = tot = 0
    for perm in itertools.permutations(ry):
        m2 = sum(perm) / n
        nu = sum((a - mx) * (b - m2) for a, b in zip(rx, perm))
        de = math.sqrt(sum((a - mx) ** 2 for a in rx) * sum((b - m2) ** 2 for b in perm))
        r2 = nu / de if de else 0.0
        tot += 1
        if abs(r2) >= abs(rho) - 1e-12:
            cnt += 1
    return rho, cnt / tot


# ---------------------------------------------------------------- study overlap

_AUTHYEAR = re.compile(r"([A-Z][A-Za-zÀ-ſ'`-]+)")


def _key_moreno(cit):
    m = _AUTHYEAR.search(cit)
    y = re.search(r"(1[89]\d\d|20[0-2]\d)", cit)
    return (m.group(1).lower(), y.group(1)) if m and y else None


def _key_jones(ref):
    m = _AUTHYEAR.search(ref)
    y = re.search(r"\((1[89]\d\d|20[0-2]\d)\)", ref)
    return (m.group(1).lower(), y.group(1)) if m and y else None


def overlap(xlsx_path, jones_xls):
    import xlrd
    mor = {k for k in (_key_moreno(r["citation"]) for r in load(xlsx_path)) if k}
    b = xlrd.open_workbook(jones_xls)
    sh = b.sheet_by_index(0)
    jon = set()
    for r in range(241, sh.nrows):
        txt = str(sh.cell_value(r, 1)).strip()
        if len(txt) > 20:
            k = _key_jones(txt)
            if k:
                jon.add(k)
    inter = mor & jon
    print(f"\nSTUDY OVERLAP (first-author surname + year keys)")
    print(f"  Moreno-Mateos 2017 primary studies (distinct keys) : {len(mor)}")
    print(f"  Jones & Schmitz 2009 reference list (distinct keys) : {len(jon)}")
    print(f"  intersection                                        : {len(inter)}"
          f"   ({100.0*len(inter)/max(len(jon),1):.1f}% of Jones,"
          f" {100.0*len(inter)/max(len(mor),1):.1f}% of Moreno-Mateos)")
    for k in sorted(inter):
        print(f"      {k[0]} {k[1]}")


# ---------------------------------------------------------------- report

def _line(label, f):
    if f is None:
        return f"{label:<26s}  (too few events/censorings to fit)"
    n, ev, cn, b, lo, hi, eta = f
    star = "*" if hi < 1.0 else (" " if lo < 1.0 < hi else "+")
    return (f"{label:<26s} {n:5d} {ev:5d} {cn:5d}   {b:5.3f}  [{lo:5.3f}, {hi:5.3f}] "
            f"{eta:9.2f}  {star}")


HDR = (f"{'group':<26s} {'N':>5s} {'ev':>5s} {'cens':>5s}   {'beta':>5s}  "
       f"{'95% profile CI':>16s} {'eta(yr)':>9s}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--likelihood", default="c29", choices=["c29", "current-status"])
    ap.add_argument("--study-level", action="store_true")
    ap.add_argument("--overlap", action="store_true")
    ap.add_argument("--jones-xls", default=None)
    args = ap.parse_args()

    rows = load(args.xlsx)
    recs, dropped = survival(rows)
    if args.study_level:
        recs = study_level(recs)

    L = args.likelihood
    print("Moreno-Mateos et al. 2017, Dryad doi:10.5061/dryad.t5c97, "
          "'Moreno, Jones database.xlsx'")
    print(f"outcome-measure rows read: {len(rows)}   usable: {len(recs)}   "
          f"likelihood: {L}   unit: {'STUDY x habitat' if args.study_level else 'outcome measure'}")
    for k, v in dropped.most_common():
        print(f"  dropped {v:4d}: {k}")
    print(f"  distinct primary studies in usable set: "
          f"{len(set(r['citation'] for r in recs))}")
    print()

    by = collections.defaultdict(list)
    for r in recs:
        by[r["habitat"]].append((r["time"], r["event"]))
    pooled = [(r["time"], r["event"]) for r in recs]

    print(HDR); print("-" * len(HDR))
    fits = {}
    for k in ("Forest", "Marine", "Freshwater", "Brackish", "Terrestrial"):
        f = fit_weibull(by[k], L)
        fits[k] = f
        print(_line(k, f))
    print(_line("ALL POOLED", fit_weibull(pooled, L)))
    print("  * = CI entirely below 1 (decreasing hazard); + = entirely above 1")

    # ------------------------------------------------ P1
    common = [k for k in C29_BETA if fits.get(k)]
    a = [C29_BETA[k] for k in common]
    b = [fits[k][3] for k in common]
    rho, p = spearman(a, b)
    print(f"\nPREDICTION 1  (C29 sec.5.1): sign of the per-habitat rank correlation")
    print(f"  habitats in common: {common}")
    print(f"  C29 beta : {[round(x,3) for x in a]}")
    print(f"  C32 beta : {[round(x,3) for x in b]}")
    print(f"  Spearman rho = {rho:+.3f}   n = {len(common)}   "
          f"exact two-sided p = {p:.3f}")
    print(f"  VERDICT: prediction 1 {'PASSES' if rho > 0 else 'FAILS'} on sign "
          f"(predicted positive).")

    # ------------------------------------------------ P2
    print(f"\nPREDICTION 2  (C29 sec.5.2): frailty split by response-variable class")
    print(HDR); print("-" * len(HDR))
    bym = collections.defaultdict(list)
    byc = collections.defaultdict(list)
    for r in recs:
        bym[r["metric"]].append((r["time"], r["event"]))
        byc[METRIC_CLASS.get(r["metric"], "?")].append((r["time"], r["event"]))
    for k in sorted(byc):
        print(_line("class: " + k, fit_weibull(byc[k], L)))
    for k in sorted(bym, key=lambda x: -len(bym[x])):
        print(_line("  metric: " + k, fit_weibull(bym[k], L)))
    print("\n  within habitat x class:")
    bhc = collections.defaultdict(list)
    for r in recs:
        bhc[(r["habitat"], METRIC_CLASS.get(r["metric"], "?"))].append(
            (r["time"], r["event"]))
    for k in sorted(bhc, key=lambda x: -len(bhc[x])):
        if len(bhc[k]) >= 25:
            print(_line(f"  {k[0]} / {k[1][:14]}", fit_weibull(bhc[k], L)))

    # ------------------------------------------------ disturbance
    print("\nby disturbance category (N >= 30):")
    print(HDR); print("-" * len(HDR))
    byd = collections.defaultdict(list)
    for r in recs:
        byd[r["disturbance"]].append((r["time"], r["event"]))
    for k, v in sorted(byd.items(), key=lambda kv: -len(kv[1])):
        if len(v) >= 30:
            print(_line(k[:26], fit_weibull(v, L)))

    if args.overlap and args.jones_xls:
        overlap(args.xlsx, args.jones_xls)
    return 0


if __name__ == "__main__":
    sys.exit(main())
